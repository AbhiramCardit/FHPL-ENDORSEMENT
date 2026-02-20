"""
ABHI XLS/XLSX Extractor — extracts endorsement data from structured spreadsheets.

Ported from the standalone endorsement-sheet-extractor script.
Handles merged cells, header metadata, tabular records, and summary sections.

Supports both .xls (via xlrd) and .xlsx (via openpyxl) formats.
"""

from __future__ import annotations

import os
from typing import Any

from app.core.logging import get_logger

logger = get_logger(__name__)


# ═══════════════════════════════════════════════════════════
#  Sheet Adapters — uniform interface over xlrd / openpyxl
# ═══════════════════════════════════════════════════════════

class XlrdSheetAdapter:
    """Adapter for xlrd sheets (0-based indexing)."""

    def __init__(self, sheet) -> None:
        self._s = sheet
        self.nrows = sheet.nrows
        self.ncols = sheet.ncols

    def raw_value(self, r: int, c: int) -> Any:
        value = self._s.cell_value(r, c)
        return value if value != "" else None

    def merged_ranges(self):
        for rlo, rhi, clo, chi in self._s.merged_cells:
            yield rlo, rhi, clo, chi


class OpenpyxlSheetAdapter:
    """Adapter for openpyxl worksheets (converts 1-based to 0-based)."""

    def __init__(self, ws) -> None:
        self._ws = ws
        self.nrows = ws.max_row
        self.ncols = ws.max_column

    def raw_value(self, r: int, c: int) -> Any:
        value = self._ws.cell(r + 1, c + 1).value
        return value if value is not None else None

    def merged_ranges(self):
        for merged_range in self._ws.merged_cells.ranges:
            yield (
                merged_range.min_row - 1,
                merged_range.max_row,
                merged_range.min_col - 1,
                merged_range.max_col,
            )


# ═══════════════════════════════════════════════════════════
#  Core Extraction Logic
# ═══════════════════════════════════════════════════════════

def _load_sheet(path: str):
    """Load the first sheet from an XLS or XLSX file."""
    extension = os.path.splitext(path)[1].lower()

    if extension == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError(
                "xlrd is required for .xls files. Install with: pip install xlrd==1.2.0"
            ) from exc
        workbook = xlrd.open_workbook(path)
        return XlrdSheetAdapter(workbook.sheet_by_index(0))

    if extension == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for .xlsx files. Install with: pip install openpyxl"
            ) from exc
        workbook = load_workbook(path)
        return OpenpyxlSheetAdapter(workbook.active)

    raise ValueError(f"Unsupported extension: {extension}")


def _build_merged_lookup(sheet) -> dict[tuple[int, int], Any]:
    """Build a lookup: (row, col) → top-left value for all merged cells."""
    lookup = {}
    for rlo, rhi, clo, chi in sheet.merged_ranges():
        top_left = sheet.raw_value(rlo, clo)
        for row in range(rlo, rhi):
            for col in range(clo, chi):
                lookup[(row, col)] = top_left
    return lookup


def _value_at(sheet, row: int, col: int, merged_lookup: dict) -> Any:
    """Get the effective value at (row, col), respecting merged cells."""
    value = sheet.raw_value(row, col)
    return value if value is not None else merged_lookup.get((row, col))


def _is_serial_value(value) -> bool:
    """Check if a value looks like a serial number (numeric)."""
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return False
        try:
            float(text)
            return True
        except ValueError:
            return False
    return False


def _header_looks_serial(header_name: str) -> bool:
    """Check if a header name refers to a serial number column."""
    normalized = "".join(ch for ch in header_name.lower() if ch.isalnum())
    serial_aliases = {"srno", "sno", "serialno", "serialnumber", "srnumber"}
    return normalized in serial_aliases or normalized.startswith("srno")


def _find_table_header_row(sheet) -> int:
    """Find the row that contains the table header (≥8 distinct values)."""
    for row in range(sheet.nrows):
        distinct_values = set()
        for col in range(sheet.ncols):
            value = sheet.raw_value(row, col)
            if value is None:
                continue
            normalized = str(value).strip()
            if normalized:
                distinct_values.add(normalized)
        if len(distinct_values) >= 8:
            return row
    return sheet.nrows - 1


def _extract_header(sheet, table_header_row: int) -> dict[str, Any]:
    """Extract key-value metadata from rows above the table header."""
    merge_top_lefts = {}
    merge_members = set()
    for rlo, rhi, clo, chi in sheet.merged_ranges():
        merge_top_lefts[(rlo, clo)] = chi
        for row in range(rlo, rhi):
            for col in range(clo, chi):
                if row == rlo and col == clo:
                    continue
                merge_members.add((row, col))

    def merged_end(row, col):
        return merge_top_lefts.get((row, col), col + 1)

    output: dict[str, Any] = {}
    title_found = False

    for row in range(table_header_row):
        entries = []
        for col in range(sheet.ncols):
            if (row, col) in merge_members:
                continue
            value = sheet.raw_value(row, col)
            if value is None:
                continue
            if not str(value).strip():
                continue
            entries.append((col, value, merged_end(row, col)))

        if not entries:
            continue

        if len(entries) == 1 and isinstance(entries[0][1], str) and not title_found:
            output["_title"] = entries[0][1].strip()
            title_found = True
            continue

        index = 0
        while index < len(entries):
            _, label_value, _ = entries[index]
            if not isinstance(label_value, str) or not label_value.strip():
                index += 1
                continue
            label = label_value.strip()
            if index + 1 < len(entries):
                _, mapped_value, _ = entries[index + 1]
                output[label] = (
                    mapped_value.strip()
                    if isinstance(mapped_value, str)
                    else mapped_value
                )
                index += 2
            else:
                index += 1

    return output


def _extract_table(
    sheet, merged_lookup: dict, table_header_row: int
) -> tuple[list[dict[str, Any]], int]:
    """Extract tabular records starting from the header row."""
    column_names = {}
    for col in range(sheet.ncols):
        value = _value_at(sheet, table_header_row, col, merged_lookup)
        if value is None:
            continue
        normalized = str(value).strip()
        if normalized:
            column_names[col] = normalized

    if not column_names:
        return [], table_header_row + 1

    ordered_cols = sorted(column_names.keys())
    first_col = ordered_cols[0]
    enforce_serial = _header_looks_serial(column_names[first_col])
    records: list[dict[str, Any]] = []
    data_started = False
    last_data_row = table_header_row

    for row_index in range(table_header_row + 1, sheet.nrows):
        row_obj = {}
        for col in ordered_cols:
            value = _value_at(sheet, row_index, col, merged_lookup)
            if isinstance(value, str):
                value = value.strip() or None
            row_obj[column_names[col]] = value

        first_value = row_obj.get(column_names[first_col])

        if all(value is None for value in row_obj.values()):
            continue

        if first_value is None:
            continue

        if enforce_serial and not _is_serial_value(first_value):
            if data_started:
                break
            continue

        records.append(row_obj)
        data_started = True
        last_data_row = row_index

    data_end_row = last_data_row + 1 if data_started else table_header_row + 1
    return records, data_end_row


def _extract_summary(
    sheet, merged_lookup: dict, data_end_row: int
) -> dict[str, Any]:
    """Extract summary key-value pairs from rows below the table."""
    summary: dict[str, Any] = {}

    for row in range(data_end_row, sheet.nrows):
        cells = []
        for col in range(sheet.ncols):
            value = _value_at(sheet, row, col, merged_lookup)
            if value is None:
                continue
            text = str(value).strip()
            if text and text != " ":
                cells.append((col, value))

        if not cells:
            continue

        label = None
        label_col_idx = None
        for index, (col, value) in enumerate(cells):
            if isinstance(value, str) and value.strip():
                label = value.strip()
                label_col_idx = index
                break

        if label is None:
            continue

        mapped_value = None
        for col, value in cells[label_col_idx + 1 :]:
            if isinstance(value, str):
                text = value.strip()
                if text == "" or text == " " or text == label:
                    continue
                mapped_value = text
            else:
                mapped_value = value
            break

        if mapped_value is None:
            continue

        summary[label] = mapped_value

    return summary


# ═══════════════════════════════════════════════════════════
#  Public API
# ═══════════════════════════════════════════════════════════

def extract_xls(path: str) -> dict[str, Any]:
    """
    Extract endorsement data from an ABHI XLS/XLSX file.

    Returns:
        dict with keys:
            - title: str | None
            - header: dict[str, Any]   — metadata above the table
            - records: list[dict]      — one dict per row
            - summary: dict[str, Any]  — totals/footer below the table

    Raises:
        FileNotFoundError: If the file doesn't exist.
        ValueError: If the file extension is unsupported.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")

    logger.info("ABHI XLS extraction starting", path=path)

    sheet = _load_sheet(path)
    merged_lookup = _build_merged_lookup(sheet)
    table_header_row = _find_table_header_row(sheet)
    header = _extract_header(sheet, table_header_row)
    records, data_end_row = _extract_table(sheet, merged_lookup, table_header_row)
    summary = _extract_summary(sheet, merged_lookup, data_end_row)

    title = header.pop("_title", None)
    result: dict[str, Any] = {}
    if title:
        result["title"] = title
    result["header"] = header
    result["records"] = records
    result["summary"] = summary

    logger.info(
        "ABHI XLS extraction complete",
        title=title,
        header_fields=len(header),
        records=len(records),
        summary_fields=len(summary),
    )

    return result
